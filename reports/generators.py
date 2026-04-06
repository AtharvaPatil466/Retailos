"""Report generators for PDF and Excel exports."""

import io
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

GST_RATES = {
    "Grocery": 0.05, "Dairy": 0.05, "Frozen": 0.12, "Snacks": 0.12,
    "Beverages": 0.12, "Personal Care": 0.18, "Cleaning": 0.18,
    "Baby Care": 0.12, "Bakery": 0.05, "Protein & Health": 0.18,
}


def generate_sales_excel(orders: list[dict], date_from: str, date_to: str) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Header styling
    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    # Title row
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Sales Report: {date_from} to {date_to}"
    ws["A1"].font = Font(bold=True, size=14)

    # Headers
    headers = ["Order ID", "Date", "Customer", "Items", "Subtotal", "GST", "Total", "Payment"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = 4
    total_revenue = 0
    total_gst = 0
    for order in orders:
        ts = order.get("timestamp", 0)
        dt = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M") if ts else ""
        items_str = ", ".join(f"{i['product_name']} x{i['qty']}" for i in order.get("items", []))
        subtotal = order.get("total_amount", 0) - order.get("gst_amount", 0)
        gst = order.get("gst_amount", 0)

        ws.cell(row=row, column=1, value=order.get("order_id", ""))
        ws.cell(row=row, column=2, value=dt)
        ws.cell(row=row, column=3, value=order.get("customer_name", "Walk-in"))
        ws.cell(row=row, column=4, value=items_str)
        ws.cell(row=row, column=5, value=round(subtotal, 2))
        ws.cell(row=row, column=6, value=round(gst, 2))
        ws.cell(row=row, column=7, value=round(order.get("total_amount", 0), 2))
        ws.cell(row=row, column=8, value=order.get("payment_method", "Cash"))

        total_revenue += order.get("total_amount", 0)
        total_gst += gst
        row += 1

    # Totals row
    row += 1
    ws.cell(row=row, column=4, value="TOTALS").font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(total_revenue - total_gst, 2)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(total_gst, 2)).font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(total_revenue, 2)).font = Font(bold=True)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_pnl_pdf(
    revenue: float,
    cost_of_goods: float,
    gst_collected: float,
    returns_amount: float,
    period: str,
    store_name: str = "RetailOS Supermart",
) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18, spaceAfter=12)
    elements.append(Paragraph(f"{store_name} — Profit & Loss Statement", title_style))
    elements.append(Paragraph(f"Period: {period}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    gross_profit = revenue - cost_of_goods
    net_revenue = revenue - returns_amount
    net_profit = net_revenue - cost_of_goods

    data = [
        ["Line Item", "Amount (INR)"],
        ["Gross Revenue", f"₹{revenue:,.2f}"],
        ["Less: Returns & Refunds", f"(₹{returns_amount:,.2f})"],
        ["Net Revenue", f"₹{net_revenue:,.2f}"],
        ["", ""],
        ["Cost of Goods Sold", f"(₹{cost_of_goods:,.2f})"],
        ["Gross Profit", f"₹{gross_profit:,.2f}"],
        ["Gross Margin", f"{(gross_profit / revenue * 100) if revenue else 0:.1f}%"],
        ["", ""],
        ["GST Collected", f"₹{gst_collected:,.2f}"],
        ["", ""],
        ["Net Profit (before tax)", f"₹{net_profit:,.2f}"],
    ]

    table = Table(data, colWidths=[280, 180])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5233")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F0")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf


def generate_gst_excel(orders: list[dict], date_from: str, date_to: str) -> io.BytesIO:
    """GST returns summary grouped by tax slab."""
    wb = Workbook()
    ws = wb.active
    ws.title = "GST Summary"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"GST Returns Summary: {date_from} to {date_to}"
    ws["A1"].font = Font(bold=True, size=14)

    # Accumulate by category
    category_totals: dict[str, dict] = {}
    for order in orders:
        for item in order.get("items", []):
            cat = item.get("category", "Grocery")
            if cat not in category_totals:
                category_totals[cat] = {"taxable_value": 0, "items_count": 0}
            category_totals[cat]["taxable_value"] += item.get("total", item.get("unit_price", 0) * item.get("qty", 1))
            category_totals[cat]["items_count"] += item.get("qty", 1)

    headers = ["Category", "GST Rate", "Taxable Value", "CGST", "SGST", "Total GST"]
    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    row = 4
    grand_taxable = 0
    grand_gst = 0
    for cat, data in sorted(category_totals.items()):
        rate = GST_RATES.get(cat, 0.12)
        taxable = data["taxable_value"]
        gst = taxable * rate
        cgst = gst / 2
        sgst = gst / 2

        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=2, value=f"{rate * 100:.0f}%")
        ws.cell(row=row, column=3, value=round(taxable, 2))
        ws.cell(row=row, column=4, value=round(cgst, 2))
        ws.cell(row=row, column=5, value=round(sgst, 2))
        ws.cell(row=row, column=6, value=round(gst, 2))

        grand_taxable += taxable
        grand_gst += gst
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=round(grand_taxable, 2)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(grand_gst, 2)).font = Font(bold=True)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_inventory_excel(products: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Inventory Report — {date.today().isoformat()}"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["SKU", "Product", "Category", "Stock", "Threshold", "Daily Rate", "Days Left", "Status"]
    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    for row_idx, p in enumerate(products, 4):
        daily = p.get("daily_sales_rate", 0)
        stock = p.get("current_stock", 0)
        days_left = stock / daily if daily > 0 else 999
        status = "Critical" if days_left < 2 else "Warning" if days_left < 5 else "OK"

        ws.cell(row=row_idx, column=1, value=p.get("sku", ""))
        ws.cell(row=row_idx, column=2, value=p.get("product_name", ""))
        ws.cell(row=row_idx, column=3, value=p.get("category", ""))
        ws.cell(row=row_idx, column=4, value=stock)
        ws.cell(row=row_idx, column=5, value=p.get("reorder_threshold", 0))
        ws.cell(row=row_idx, column=6, value=daily)
        ws.cell(row=row_idx, column=7, value=round(days_left, 1))
        cell = ws.cell(row=row_idx, column=8, value=status)
        if status == "Critical":
            cell.font = Font(color="CC0000", bold=True)
        elif status == "Warning":
            cell.font = Font(color="CC8800")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_inventory_pdf(products: list[dict]) -> io.BytesIO:
    """Generate a PDF inventory report with stock status indicators."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18, spaceAfter=12)
    elements.append(Paragraph(f"Inventory Report — {date.today().isoformat()}", title_style))
    elements.append(Spacer(1, 12))

    total_items = len(products)
    critical = sum(1 for p in products if _days_left(p) < 2)
    warning = sum(1 for p in products if 2 <= _days_left(p) < 5)
    total_value = sum(p.get("current_stock", 0) * p.get("unit_price", 0) for p in products)

    summary_data = [
        ["Total SKUs", str(total_items)],
        ["Critical (< 2 days)", str(critical)],
        ["Warning (< 5 days)", str(warning)],
        ["Total Stock Value", f"Rs.{total_value:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[200, 200])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E8F5E9")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    data = [["SKU", "Product", "Stock", "Threshold", "Days Left", "Status"]]
    for p in sorted(products, key=lambda x: _days_left(x)):
        days = _days_left(p)
        status = "CRITICAL" if days < 2 else "WARNING" if days < 5 else "OK"
        data.append([
            p.get("sku", ""),
            p.get("product_name", "")[:30],
            str(p.get("current_stock", 0)),
            str(p.get("reorder_threshold", 0)),
            f"{days:.0f}" if days < 999 else "-",
            status,
        ])

    table = Table(data, colWidths=[70, 150, 50, 60, 55, 60])
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5233")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F0")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for i, row in enumerate(data[1:], 1):
        if row[-1] == "CRITICAL":
            style_cmds.append(("TEXTCOLOR", (-1, i), (-1, i), colors.HexColor("#CC0000")))
        elif row[-1] == "WARNING":
            style_cmds.append(("TEXTCOLOR", (-1, i), (-1, i), colors.HexColor("#CC8800")))

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf


def generate_customer_excel(customers: list[dict], date_from: str = "", date_to: str = "") -> io.BytesIO:
    """Generate customer analytics Excel report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Report"

    period = f"{date_from} to {date_to}" if date_from else date.today().isoformat()
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Customer Report — {period}"
    ws["A1"].font = Font(bold=True, size=14)

    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    headers = ["Customer ID", "Name", "Phone", "Total Orders", "Total Spent", "Loyalty Tier", "Outstanding Credit"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    for row_idx, c in enumerate(customers, 4):
        ws.cell(row=row_idx, column=1, value=c.get("customer_code", ""))
        ws.cell(row=row_idx, column=2, value=c.get("name", ""))
        ws.cell(row=row_idx, column=3, value=c.get("phone", ""))
        ws.cell(row=row_idx, column=4, value=c.get("total_orders", 0))
        ws.cell(row=row_idx, column=5, value=round(c.get("total_spent", 0), 2))
        ws.cell(row=row_idx, column=6, value=c.get("loyalty_tier", "bronze"))
        ws.cell(row=row_idx, column=7, value=round(c.get("outstanding_credit", 0), 2))

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_daily_summary_pdf(
    date_str: str,
    revenue: float,
    orders_count: int,
    top_products: list[dict],
    payment_breakdown: dict[str, float],
    store_name: str = "RetailOS Supermart",
) -> io.BytesIO:
    """Generate a single-day summary PDF."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=18, spaceAfter=12)
    elements.append(Paragraph(f"{store_name} — Daily Summary", title_style))
    elements.append(Paragraph(f"Date: {date_str}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    avg_order = revenue / orders_count if orders_count else 0
    kpi_data = [
        ["Metric", "Value"],
        ["Total Revenue", f"Rs.{revenue:,.2f}"],
        ["Total Orders", str(orders_count)],
        ["Average Order Value", f"Rs.{avg_order:,.2f}"],
    ]
    kpi_table = Table(kpi_data, colWidths=[200, 200])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5233")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 16))

    if payment_breakdown:
        elements.append(Paragraph("Payment Methods", styles["Heading2"]))
        pay_data = [["Method", "Amount"]]
        for method, amount in payment_breakdown.items():
            pay_data.append([method, f"Rs.{amount:,.2f}"])
        pay_table = Table(pay_data, colWidths=[200, 200])
        pay_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a365d")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(pay_table)
        elements.append(Spacer(1, 16))

    if top_products:
        elements.append(Paragraph("Top Products", styles["Heading2"]))
        prod_data = [["Product", "Qty Sold", "Revenue"]]
        for p in top_products[:10]:
            prod_data.append([p.get("name", ""), str(p.get("qty_sold", 0)), f"Rs.{p.get('revenue', 0):,.2f}"])
        prod_table = Table(prod_data, colWidths=[200, 80, 120])
        prod_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#92400e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DDDDDD")),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(prod_table)

    doc.build(elements)
    buf.seek(0)
    return buf


def _days_left(product: dict) -> float:
    daily = product.get("daily_sales_rate", 0)
    stock = product.get("current_stock", 0)
    return stock / daily if daily > 0 else 999
