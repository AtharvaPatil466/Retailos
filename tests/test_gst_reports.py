"""Tests for GST return generation and P&L reports."""

import io
from openpyxl import load_workbook

from reports.gst_returns import (
    generate_gstr1_excel,
    generate_gstr3b_excel,
    generate_pnl_excel,
)


def _make_invoices():
    return [
        {
            "invoice_number": "INV-001",
            "invoice_date": "2024-10-15",
            "buyer_gstin": "27AABCU9603R1ZM",
            "total_amount": 3650.0,
            "gst_rate": 5,
            "taxable_value": 3476.19,
            "place_of_supply": "27-Maharashtra",
            "items": [
                {"hsn_code": "1006", "description": "Basmati Rice 5kg", "qty": 10, "total": 2750.0},
                {"hsn_code": "1512", "description": "Sunflower Oil 1L", "qty": 5, "total": 900.0},
            ],
        },
        {
            "invoice_number": "INV-002",
            "invoice_date": "2024-10-18",
            "buyer_gstin": "",
            "total_amount": 135.0,
            "gst_rate": 18,
            "taxable_value": 114.41,
            "items": [
                {"hsn_code": "3401", "description": "Soap Bar 100g", "qty": 3, "total": 135.0},
            ],
        },
    ]


def test_gstr1_excel_generates():
    invoices = _make_invoices()
    buf = generate_gstr1_excel(invoices, "2024-10-01", "2024-10-31", gstin="27AABCU9603R1ZM")
    assert isinstance(buf, io.BytesIO)
    wb = load_workbook(buf)
    assert "B2B" in wb.sheetnames
    assert "B2C" in wb.sheetnames
    assert "HSN Summary" in wb.sheetnames


def test_gstr1_b2b_has_gstin_entries():
    invoices = _make_invoices()
    buf = generate_gstr1_excel(invoices, "2024-10-01", "2024-10-31", gstin="27AABCU9603R1ZM")
    wb = load_workbook(buf)
    b2b = wb["B2B"]
    # Header rows + at least 1 data row (invoice with GSTIN)
    assert b2b.max_row >= 5


def test_gstr1_b2c_has_non_gstin_entries():
    invoices = _make_invoices()
    buf = generate_gstr1_excel(invoices, "2024-10-01", "2024-10-31", gstin="27AABCU9603R1ZM")
    wb = load_workbook(buf)
    b2c = wb["B2C"]
    assert b2c.max_row >= 4


def test_gstr3b_excel_generates():
    sales_data = {"taxable_value": 3590.60, "gst_collected": 200.0}
    purchase_data = {"gst_paid": 125.0}
    buf = generate_gstr3b_excel(sales_data, purchase_data, "2024-10-01", "2024-10-31")
    assert isinstance(buf, io.BytesIO)
    wb = load_workbook(buf)
    assert "GSTR-3B" in wb.sheetnames


def test_pnl_excel_generates():
    buf = generate_pnl_excel(
        revenue=150000.0,
        cost_of_goods=90000.0,
        gst_collected=12000.0,
        returns_amount=5000.0,
        expenses={"Rent": 15000.0, "Salaries": 25000.0, "Utilities": 3000.0, "Marketing": 2000.0},
        period="October 2024",
    )
    assert isinstance(buf, io.BytesIO)
    wb = load_workbook(buf)
    assert "Profit & Loss" in wb.sheetnames


def test_gstr1_empty_orders():
    buf = generate_gstr1_excel([], "2024-10-01", "2024-10-31", gstin="27AABCU9603R1ZM")
    assert isinstance(buf, io.BytesIO)
    wb = load_workbook(buf)
    assert "B2B" in wb.sheetnames
